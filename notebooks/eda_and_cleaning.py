import os
import json
import openpyxl

# Definimos las rutas relativas basadas en la localización del script
script_dir = os.path.dirname(os.path.abspath(__file__)) 
project_root = os.path.dirname(script_dir)
excel_path = os.path.join(project_root, "data", "raw", "India_dataset.xlsx")
output_path = os.path.join(project_root, "data", "processed", "india_data.json")

print(f"Reading raw data from: {excel_path}")
wb = openpyxl.load_workbook(excel_path, data_only=True)

# Procesamos los metadatos
metadata_sheet = wb['Series - Metadata']
meta_rows = list(metadata_sheet.iter_rows(values_only=True))
headers_meta = meta_rows[0]

indicators_meta = {}
for row in meta_rows[1:]:
    if not row or not row[0]:
        continue
    code = row[0]
    indicators_meta[code] = {
        "code": code,
        "license": row[1],
        "name": row[2],
        "definition": row[3],
        "source": row[4],
        "topic": row[5],
        "dataset": row[6],
        "unit": row[7]
    }

# Procesamos los datos
data_sheet = wb['Data']
data_rows = list(data_sheet.iter_rows(values_only=True))
headers_data = data_rows[0]

# Extraemos los años correspondientes
year_cols = []
for h in headers_data[4:]:
    if h is None:
        continue
    h_str = str(h).strip()
    if not h_str:
        continue
    if ' [YR' in h_str:
        year_str = h_str.split(' [YR')[0]
        year_cols.append(int(year_str))
    else:
        year_cols.append(h_str)

parsed_data = []
for row_idx, row in enumerate(data_rows[1:]):
    if not row or not row[0]:
        continue
    
    country_name = row[0]
    if str(country_name).startswith('Data from database'):
        continue
    if str(country_name).startswith('Last Updated'):
        continue
        
    country_code = row[1]
    series_name = row[2]
    series_code = row[3]
    
    if not series_code:
        continue

    yearly_values = {}
    data_values = row[4:]
    for idx in range(min(len(year_cols), len(data_values))):
        year = year_cols[idx]
        val = data_values[idx]
        
        # Limpiamos y normalizamos los valores
        if val == '..' or val is None:
            cleaned_val = None
        else:
            try:
                cleaned_val = float(val)
            except ValueError:
                cleaned_val = val
        yearly_values[str(year)] = cleaned_val
        
    meta = indicators_meta.get(series_code, {})
    
    parsed_data.append({
        "series_code": series_code,
        "series_name": series_name,
        "unit": meta.get("unit", "") or "unspecified",
        "topic": meta.get("topic", "") or "unspecified",
        "definition": meta.get("definition", "") or "No definition available.",
        "source": meta.get("source", "") or "World Bank (WDI)",
        "values": yearly_values
    })

output_data = {
    "country_name": "India",
    "country_code": "IND",
    "indicators": parsed_data,
    "years": [str(y) for y in year_cols]
}

# Creamos el directorio de salida
os.makedirs(os.path.dirname(output_path), exist_ok=True)

with open(output_path, 'w', encoding='utf-8') as f:
    json.dump(output_data, f, indent=2, ensure_ascii=False)

print(f"Data successfully cleaned and saved to {output_path}!")
print(f"Number of indicators processed: {len(parsed_data)}")

# Mostramos el resumen de cobertura
print("\n--- Summary of Indicators Coverage ---")
for ind in parsed_data:
    non_null_years = [y for y, v in ind["values"].items() if v is not None]
    coverage = len(non_null_years)
    print(f"Code: {ind['series_code']}")
    print(f"  Name: {ind['series_name']}")
    print(f"  Coverage: {coverage}/{len(year_cols)} years ({coverage/len(year_cols)*100:.1f}%)")
    if coverage > 0:
        print(f"  Range: {min(non_null_years)} - {max(non_null_years)}")
        # Calculamos las estadísticas descriptivas
        vals = [v for v in ind["values"].values() if isinstance(v, (int, float))]
        if vals:
            print(f"  Stats: Min={min(vals):.3f}, Max={max(vals):.3f}, Mean={sum(vals)/len(vals):.3f}")
    print()
